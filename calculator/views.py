from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib import messages
from .forms import SalaryCalculationForm
from .models import SalaryCalculation
import openpyxl
from openpyxl.styles import Font, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from io import BytesIO
import datetime

def calculate_salary(data):
    """Calculate salary based on Moroccan tax system"""
    salaire_base = data['salaire_base']
    primes_conges = data['primes_conges']
    indemnite_exoneree = data['indemnite_exoneree']
    nb_charge_familiale = data['nb_charge_familiale']
    
    # 1. Salaire brut
    salaire_brut = salaire_base + primes_conges + indemnite_exoneree
    salaire_brut_imposable = salaire_base + primes_conges

    # 2. Cotisations sociales
    cnss = min(salaire_brut_imposable, 6000) * 0.0448
    amo = salaire_brut_imposable * 0.0226
    mutuelle = salaire_brut_imposable * 0.025

    # 3. Frais professionnels
    if salaire_brut_imposable < 6500:
        frais_pro = salaire_brut_imposable * 0.35
    else:
        frais_pro = salaire_brut_imposable * 0.25
    frais_pro = min(frais_pro, 2916.66)

    # 4. Salaire Net Imposable
    sni = salaire_brut_imposable - cnss - amo - mutuelle - frais_pro

    # 5. Calcul IR (nouveau barème)
    if sni <= 3333:
        taux, somme_deduire = 0.00, 0.00
    elif sni <= 5000:
        taux, somme_deduire = 0.10, 333.33
    elif sni <= 6667:
        taux, somme_deduire = 0.20, 833.33
    elif sni <= 8333:
        taux, somme_deduire = 0.30, 1500.00
    elif sni <= 15000:
        taux, somme_deduire = 0.34, 1833.33
    else:
        taux, somme_deduire = 0.37, 2283.33

    ir_brut = max(0, sni * taux - somme_deduire)

    # 6. Charge famille (max 6 personnes)
    charge_famille = min(nb_charge_familiale, 6) * 41.67

    # 7. IR net et salaire net
    ir_net = max(0, ir_brut - charge_famille)
    salaire_net = salaire_brut - cnss - amo - mutuelle - ir_net

    return {
        'salaire_brut': round(salaire_brut, 2),
        'salaire_brut_imposable': round(salaire_brut_imposable, 2),
        'cnss': round(cnss, 2),
        'amo': round(amo, 2),
        'mutuelle': round(mutuelle, 2),
        'frais_professionnels': round(frais_pro, 2),
        'salaire_net_imposable': round(sni, 2),
        'ir_brut': round(ir_brut, 2),
        'charge_famille': round(charge_famille, 2),
        'ir_net': round(ir_net, 2),
        'salaire_net': round(salaire_net, 2)
    }


def home(request):
    if request.method == 'POST':
        form = SalaryCalculationForm(request.POST)
        if form.is_valid():
            # Get form data
            form_data = form.cleaned_data
            
            # Calculate salary
            calculations = calculate_salary(form_data)
            
            # Create database record
            calculation = SalaryCalculation.objects.create(
                salaire_base=form_data['salaire_base'],
                primes_conges=form_data['primes_conges'],
                indemnite_exoneree=form_data['indemnite_exoneree'],
                nb_charge_familiale=form_data['nb_charge_familiale'],
                **calculations
            )
            
            messages.success(request, 'Calcul effectué avec succès!')
            return redirect('result', calculation_id=calculation.id)
    else:
        form = SalaryCalculationForm()
    
    return render(request, 'calculator/home.html', {'form': form})

def result(request, calculation_id):
    try:
        calculation = SalaryCalculation.objects.get(id=calculation_id)
        return render(request, 'calculator/result.html', {'calculation': calculation})
    except SalaryCalculation.DoesNotExist:
        messages.error(request, 'Calcul non trouvé.')
        return redirect('home')

def history(request):
    calculations = SalaryCalculation.objects.all()[:20]  # Last 20 calculations
    return render(request, 'calculator/history.html', {'calculations': calculations})

from django.http import HttpResponse
from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .models import SalaryCalculation


def export_excel(request, calculation_id):
    try:
        calculation = SalaryCalculation.objects.get(id=calculation_id)
    except SalaryCalculation.DoesNotExist:
        messages.error(request, 'Calcul non trouvé.')
        return redirect('home')
    
    # Créer le classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Calcul Salaire"
    
    # En-tête
    ws['A1'] = "CALCUL DE SALAIRE"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    
    # Données d'entrée
    ws['A3'] = "DONNÉES D'ENTRÉE"
    ws['A3'].font = Font(bold=True)
    ws['A4'] = "Salaire de Base"
    ws['B4'] = calculation.salaire_base
    ws['A5'] = "Primes et Congés"
    ws['B5'] = calculation.primes_conges
    ws['A6'] = "Indemnité Exonérée"
    ws['B6'] = calculation.indemnite_exoneree
    ws['A7'] = "Charges Familiales"
    ws['B7'] = calculation.nb_charge_familiale
    
    # Résultats
    ws['A9'] = "RÉSULTATS DU CALCUL"
    ws['A9'].font = Font(bold=True)
    ws['A10'] = "Salaire Brut"
    ws['B10'] = calculation.salaire_brut
    ws['A11'] = "Salaire Brut Imposable"
    ws['B11'] = calculation.salaire_brut_imposable
    ws['A12'] = "CNSS"
    ws['B12'] = calculation.cnss
    ws['A13'] = "AMO"
    ws['B13'] = calculation.amo
    ws['A14'] = "Mutuelle"
    ws['B14'] = calculation.mutuelle
    ws['A15'] = "Frais Professionnels"
    ws['B15'] = calculation.frais_professionnels
    ws['A16'] = "Salaire Net Imposable"
    ws['B16'] = calculation.salaire_net_imposable
    ws['A17'] = "IR Brut"
    ws['B17'] = calculation.ir_brut
    ws['A18'] = "Charge Famille"
    ws['B18'] = calculation.charge_famille
    ws['A19'] = "IR Net"
    ws['B19'] = calculation.ir_net
    ws['A20'] = "SALAIRE NET"
    ws['B20'] = calculation.salaire_net
    ws['A20'].font = Font(bold=True)
    ws['B20'].font = Font(bold=True)

    # Ajustement automatique des colonnes (sans erreur)
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="calcul_salaire_{calculation.id}.xlsx"'
    
    wb.save(response)
    return response

def export_pdf(request, calculation_id):
    try:
        calculation = SalaryCalculation.objects.get(id=calculation_id)
    except SalaryCalculation.DoesNotExist:
        messages.error(request, 'Calcul non trouvé.')
        return redirect('home')
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title = Paragraph("CALCUL DE SALAIRE", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Input data table
    input_data = [
        ['DONNÉES D\'ENTRÉE', ''],
        ['Salaire de Base', f"{calculation.salaire_base} MAD"],
        ['Primes et Congés', f"{calculation.primes_conges} MAD"],
        ['Indemnité Exonérée', f"{calculation.indemnite_exoneree} MAD"],
        ['Charges Familiales', str(calculation.nb_charge_familiale)],
    ]
    
    input_table = Table(input_data, colWidths=[3*72, 2*72])
    input_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(input_table)
    story.append(Spacer(1, 20))
    
    # Results table
    results_data = [
        ['RÉSULTATS DU CALCUL', ''],
        ['Salaire Brut', f"{calculation.salaire_brut} MAD"],
        ['Salaire Brut Imposable', f"{calculation.salaire_brut_imposable} MAD"],
        ['CNSS', f"{calculation.cnss} MAD"],
        ['AMO', f"{calculation.amo} MAD"],
        ['Mutuelle', f"{calculation.mutuelle} MAD"],
        ['Frais Professionnels', f"{calculation.frais_professionnels} MAD"],
        ['Salaire Net Imposable', f"{calculation.salaire_net_imposable} MAD"],
        ['IR Brut', f"{calculation.ir_brut} MAD"],
        ['Charge Famille', f"{calculation.charge_famille} MAD"],
        ['IR Net', f"{calculation.ir_net} MAD"],
        ['SALAIRE NET', f"{calculation.salaire_net} MAD"],
    ]
    
    results_table = Table(results_data, colWidths=[3*72, 2*72])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgreen),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(results_table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="calcul_salaire_{calculation.id}.pdf"'
    response.write(buffer.getvalue())
    buffer.close()
    
    return response